"""Deterministic PII masking before customer data reaches storage or an LLM."""

import csv
import hashlib
import hmac
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path

from config import get_settings

MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100

_EMAIL_RE = re.compile(r"(?<![\w.+-])([\w.+-]+)@([\w.-]+\.[A-Za-z]{2,})(?![\w.-])")
_PHONE_RE = re.compile(r"(?<!\d)(?:\+?84|0)(?:[ .-]?\d){9,10}(?!\d)")
_CARD_RE = re.compile(r"(?<!\d)(?:\d[ -]?){13,19}(?!\d)")
_NATIONAL_ID_RE = re.compile(r"(?<!\d)(?:\d{9}|\d{12})(?!\d)")

_SENSITIVE_COLUMN_TOKENS = frozenset(
    {
        "name",
        "fullname",
        "firstname",
        "lastname",
        "hoten",
        "tenkhachhang",
        "email",
        "mail",
        "phone",
        "mobile",
        "telephone",
        "sodienthoai",
        "cccd",
        "cmnd",
        "nationalid",
        "identitynumber",
        "passport",
        "creditcard",
        "cardnumber",
        "bankaccount",
        "accountnumber",
        "address",
        "diachi",
    }
)


@dataclass(frozen=True)
class MaskedFile:
    content: bytes
    masked_columns: tuple[str, ...]


def _normalize_identifier(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    ascii_text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]", "", ascii_text.lower())


def is_sensitive_column(column_name: object) -> bool:
    normalized = _normalize_identifier(column_name)
    return any(token in normalized for token in _SENSITIVE_COLUMN_TOKENS)


def _stable_marker(value: object) -> str:
    digest = hmac.new(
        get_settings().secret_key.encode("utf-8"),
        str(value).encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()[:10]
    return f"[MASKED_{digest}]"


def mask_value(value: object, *, sensitive_column: bool = False) -> object:
    if value is None or value == "":
        return value
    if sensitive_column:
        return _stable_marker(value)
    if not isinstance(value, str):
        return value
    return mask_text(value)


def mask_text(text: str) -> str:
    """Mask PII patterns in free-form text without retaining the original value."""
    masked = _EMAIL_RE.sub(lambda match: f"masked_{_stable_marker(match.group(0))[8:-1]}@example.invalid", text)
    masked = _PHONE_RE.sub("[MASKED_PHONE]", masked)
    masked = _CARD_RE.sub("[MASKED_NUMBER]", masked)
    masked = _NATIONAL_ID_RE.sub("[MASKED_ID]", masked)
    return masked


def mask_uploaded_file(filename: str, content: bytes) -> MaskedFile:
    extension = Path(filename).suffix.lower()
    if extension in {".csv", ".tsv"}:
        return _mask_delimited(content, delimiter="\t" if extension == ".tsv" else None)
    if extension == ".xlsx":
        return _mask_xlsx(content)
    raise ValueError("Chỉ hỗ trợ che dữ liệu cho tệp CSV, TSV và XLSX.")


def _mask_delimited(content: bytes, delimiter: str | None) -> MaskedFile:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("Tệp phân tách phải dùng mã hóa UTF-8.") from exc
    sample = text[:8192]
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        return MaskedFile(content=content, masked_columns=())
    sensitive_indexes = {index for index, name in enumerate(rows[0]) if is_sensitive_column(name)}
    for row_index, row in enumerate(rows):
        if row_index == 0:
            continue
        for column_index, value in enumerate(row):
            row[column_index] = str(mask_value(value, sensitive_column=column_index in sensitive_indexes))
    output = io.StringIO(newline="")
    csv.writer(output, delimiter=delimiter, lineterminator="\n").writerows(rows)
    return MaskedFile(
        content=output.getvalue().encode("utf-8"),
        masked_columns=tuple(rows[0][index] for index in sorted(sensitive_indexes)),
    )


def _mask_xlsx(content: bytes) -> MaskedFile:
    from openpyxl import load_workbook

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            total_size = 0
            for member in archive.infolist():
                total_size += member.file_size
                if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("Tệp XLSX giải nén vượt quá giới hạn an toàn.")
                if member.file_size and (
                    member.compress_size == 0
                    or member.file_size / member.compress_size > MAX_XLSX_COMPRESSION_RATIO
                ):
                    raise ValueError("Tệp XLSX có tỷ lệ nén không an toàn.")
        workbook = load_workbook(io.BytesIO(content), keep_links=False)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("Tệp XLSX không hợp lệ hoặc đã bị hỏng.") from exc
    masked_columns: set[str] = set()
    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        sensitive_indexes = {index for index, header in enumerate(headers, start=1) if is_sensitive_column(header)}
        masked_columns.update(str(headers[index - 1]) for index in sensitive_indexes)
        for row in worksheet.iter_rows(min_row=2):
            for index, cell in enumerate(row, start=1):
                cell.value = mask_value(cell.value, sensitive_column=index in sensitive_indexes)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return MaskedFile(content=output.getvalue(), masked_columns=tuple(sorted(masked_columns)))
